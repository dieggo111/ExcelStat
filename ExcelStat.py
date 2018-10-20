#pylint: disable=W0611,E0401,C0413,R0913,W1201,C0103,R0902
import sys
import os
import logging
import re
import datetime
import yaml
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib


class ExcelStat(object):
    """
    """
    def __init__(self, folder):
        self.log = logging.getLogger(__class__.__name__)
        self.log.setLevel(logging.DEBUG)
        format_string = '%(asctime)s - %(levelname)s - %(name)s - %(message)s'
        formatter = logging.Formatter(format_string)
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        self.log.addHandler(console_handler)

        self.folder = folder
        self.sheet = "Tabelle1"
        self.files = [(item, get_file_date(item)) for item in os.listdir(folder) if not "~" in item]
        self.files = sort_files(self.files)
        self.log.info("Found files: " + str(self.files))


    def get_data_all(self, opt=None):
        data_lst = []
        for item in self.files:
            data, start, end = self.get_data_xls(item)
            i = 0
            entries = len(data[start])
            total_time = (end - start).days
            steps = total_time/entries
            for dic in data[start]:
                if dic["date"] == "No date":
                    dic["date"] = start + datetime.timedelta(days=int(steps*i))
                    i += 1
            data_lst += data[start]
            if opt is not None and "-s" in opt:
                self.get_statistics(data, start, end)
        if opt is not None and "-l" in opt:
            for dic in data_lst:
                print("{:5} {:40} {:10}".format(dic["amount"],
                                                dic["info"],
                                                str(dic["date"].date())))


    def get_data_xls(self, xls):
        file_name = os.path.join(self.folder, xls)
        self.log.info("File name: " + file_name)
        file_date = get_file_date(xls)
        self.log.info("File date: " + str(file_date.date()))

        xls_data = load_workbook(file_name)

        lst_m, lst_v, end_date = get_raw_data(xls_data,
                                              self.sheet,
                                              file_date)
        self.log.info("Entries for Marius: " + str(len(lst_m)))
        self.log.info("Entries for Viki: " + str(len(lst_v)))

        data_dict = get_data_dict(file_date, lst_m, lst_v)

        return data_dict, file_date, end_date

    def get_statistics(self, dic, start, end):
        total_time = (end- start).days
        total_m = 0
        total_v = 0
        self.log.info("End date: " + str(end.date()))
        self.log.info("Absolute time: " + str(total_time) + " days")
        for sub_dic in list(dic.values())[0]:
            if sub_dic["subject"] == "Marius":
                total_m += sub_dic["amount"]
        self.log.info("Total amount (Marius): " + str(total_m))
        for sub_dic in list(dic.values())[0]:
            if sub_dic["subject"] == "Viki":
                total_v += sub_dic["amount"]
        self.log.info("Total amount (Viki): " + str(total_v))
        self.log.info("Total amount (All): " + str(total_m + total_v))
        self.log.info("Amount per day: " + str(round((total_m + total_v)/total_time, 2)))
        print("###############################################################")


    # def plot_gen_trend(self):


def get_raw_data(xls, sheet, file_date):
    data = [[], [], [], []]
    for i, col in enumerate(xls[sheet].iter_cols(max_col=4)):
        for j, cell in enumerate(col):
            data[i].append(cell.value)

    end_date = "Not found..."
    for item in data[0]:
        if isinstance(item, str) and "Bezahlt am " in item:
            end_date = adjust_date_str(test_string(\
                    item.replace("Bezahlt am ", "")), file_date)

    lst_m = declutter_data(data[0], data[1])
    lst_v = declutter_data(data[2], data[3])

    return lst_m, lst_v, end_date

def get_data_dict(date_key, lst_m, lst_v):
    data_dict = {date_key: []}

    for val, info in lst_m:
        data_dict[date_key].append({"subject": "Marius",
                                    "amount": val,
                                    "date": adjust_date_str(test_string(info), date_key),
                                    "info": find_info(info)})
    for val, info in lst_v:
        data_dict[date_key].append({"subject": "Viki",
                                    "amount": val,
                                    "date": adjust_date_str(test_string(info), date_key),
                                    "info": find_info(info)})
    return data_dict

def declutter_data(lst1, lst2):
    new_lst = []
    for cell1, cell2 in zip(lst1, lst2):
        if isinstance(cell1, (int, float)):
            new_lst.append((cell1, cell2))
    return new_lst

def adjust_date_str(date, date_key):
    if date is None:
        return "No date"
    if date[1] == ".":
        date = "0" + date
    if len(date) == 4:
        date += "."
    if date[4] == ".":
        date = date[:3] + "0" + date[3:]
    if len(date) == 5:
        date += ("." + str(date_key.year)[2:])
    if len(date) == 6:
        date += str(date_key.year)[2:]
    if len(date) == 10:
        date = date.split("20")[0] + date.split("20")[1]
    date = datetime.datetime.strptime(date, "%d.%m.%y")
    return date

def find_info(val):
    if val is None:
        return "No Info"
    return val

def get_file_date(string):
    file_date = datetime.datetime.strptime(\
        string.replace(" ", "").replace(".xlsx", "")\
        .replace("Einkaufslisteabdem", ""), "%d.%m.%y")
    return file_date

def test_string(string):
    func_lst = [test_string_xx_xx_xxxx,
                test_string_xx_xx_xx,
                test_string_x_xx_xx,
                test_string_x_x_xx,
                test_string_xx_xx,
                test_string_xx_x,
                test_string_x_xx,
                test_string_x_x]
    if isinstance(string, str):
        for func in func_lst:
            if func(string) is not None:
                return func(string).group()
    return None

def test_string_xx_xx_xxxx(string):
    return re.search(r'\d{2}.\d{2}.\d{4}', string)
def test_string_xx_xx_xx(string):
    return re.search(r'\d{2}.\d{2}.\d{2}', string)
def test_string_x_xx_xx(string):
    return re.search(r'\d{1}.\d{2}.\d{2}', string)
def test_string_x_x_xx(string):
    return re.search(r'\d{1}.\d{1}.\d{2}', string)
def test_string_xx_xx(string):
    return re.search(r'\d{2}.\d{2}', string)
def test_string_x_xx(string):
    return re.search(r'\d{1}.\d{2}', string)
def test_string_xx_x(string):
    return re.search(r'\d{2}.\d{1}', string)
def test_string_x_x(string):
    return re.search(r'\d{1}.\d{1}', string)

def sort_files(lst):
    new_lst = []
    while True:
        if lst == []:
            break
        dates = [tup[1] for tup in lst]
        min_date = min(dates)
        min_index = dates.index(min_date)
        new_lst.append(lst[min_index][0])
        lst.pop(min_index)
    return new_lst
